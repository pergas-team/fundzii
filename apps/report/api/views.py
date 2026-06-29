from rest_framework.generics import ListCreateAPIView, RetrieveUpdateDestroyAPIView
from apps.account.permissions import AccessLevelPermission, query_set_filter_key
from django.db.models import Sum, Count, Q
import datetime
import pandas as pd
from django.http import HttpResponse
from django.utils.timezone import make_aware
from apps.lab.models import Laboratory, Request, Experiment, Device
from rest_framework.permissions import IsAuthenticated
from rest_framework.exceptions import ValidationError
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.conf import settings
import os
from .functions import generate_excel_report
from apps.account.models import User
from django.db.models import Sum, Avg
import datetime
import pandas as pd
from django.http import HttpResponse
from django.utils.timezone import make_aware
from apps.lab.models import Laboratory, Request, RequestResult
from rest_framework.permissions import IsAuthenticated
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.conf import settings
import os
import jdatetime
from decimal import Decimal
from openpyxl.utils import get_column_letter
from .functions import generate_excel_report

class ExcelReportAPIView(APIView):
    """
    API endpoint to generate and provide a link to the Excel report.
    """

    def get(self, request, format=None):
        try:
            # Generate the report
            report_path = generate_excel_report()

            # Construct the full URL for the file
            relative_file_path = os.path.relpath(report_path, settings.MEDIA_ROOT)
            file_url = request.build_absolute_uri(
                os.path.join(settings.MEDIA_URL, relative_file_path).replace('\\', '/')
            )

            return Response(
                {
                    "status": "success",
                    "data": {"url": file_url},
                    "message": "Request successful"
                },
                status=status.HTTP_200_OK
            )
        except Exception as e:
            return Response(
                {
                    "status": "error",
                    "errors": {"error": str(e)},
                    "message": "Request failed"
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class LaboratoryExcelReportAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        try:
            lab_id = request.query_params.get('lab_id')
            start_date = request.query_params.get('start_date')
            end_date = request.query_params.get('end_date')

            try:
                if start_date:
                    start_date = make_aware(datetime.datetime.strptime(start_date, '%Y-%m-%d'))
                if end_date:
                    end_date = make_aware(datetime.datetime.strptime(end_date, '%Y-%m-%d') + datetime.timedelta(days=1))
            except Exception:
                return Response(
                    {'error': 'Invalid date format. Use YYYY-MM-DD'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            labs = Laboratory.objects.all()
            if lab_id:
                labs = labs.filter(id=lab_id)

            data = []
            for lab in labs:
                requests = Request.objects.filter(
                    experiment__laboratory=lab,
                    is_completed=True,
                    parent_request__isnull=True,
                    request_status__step__id__exact=8
                ).distinct()

                if start_date:
                    requests = requests.filter(created_at__gte=start_date)
                if end_date:
                    requests = requests.filter(created_at__lt=end_date)

                total_income = requests.aggregate(total=Sum('price'))['total'] or 0
                total_income_wod = requests.aggregate(total=Sum('price_wod'))['total'] or 0
                total_grant_request_discount = requests.aggregate(total=Sum('grant_request_discount'))['total'] or 0
                total_labsnet_discount = requests.aggregate(total=Sum('labsnet_discount'))['total'] or 0
                total_request = requests.filter(parent_request__isnull=True).count()
                total_samples = sum(req.get_child_form_responses_count() for req in requests)
                operators = ', '.join([operator.get_full_name() for operator in lab.operators.all()])

                data.append({
                    'نام آزمایشگاه': lab.name,
                    'مدیر فنی': str(lab.technical_manager.get_full_name()) if lab.technical_manager else '',
                    'اپراتور': str(operators),
                    'تعداد درخواست': total_request,
                    'تعداد نمونه': total_samples,
                    'درآمد ناخالص': total_income_wod,
                    'درآمد': total_income,
                    'لبزنت': total_labsnet_discount,
                    'پژوهشی': total_grant_request_discount,
                })

            df = pd.DataFrame(data)
            df.replace('', pd.NA, inplace=True)
            df.dropna(axis=1, how='all', inplace=True)
            zero_cols = [col for col in df.select_dtypes(include='number').columns if (df[col].fillna(0) == 0).all()]
            df.drop(columns=zero_cols, inplace=True)
            now_str = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f'laboratory_report_{now_str}.xlsx'
            full_path = os.path.join(settings.MEDIA_ROOT, 'reports', filename)
            os.makedirs(os.path.dirname(full_path), exist_ok=True)

            with pd.ExcelWriter(full_path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='گزارش آزمایشگاه‌ها')

            relative_path = os.path.relpath(full_path, settings.MEDIA_ROOT)
            file_url = request.build_absolute_uri(
                os.path.join(settings.MEDIA_URL, relative_path).replace('\\', '/')
            )

            return Response(
                {
                    "status": "success",
                    "data": {"url": file_url},
                    "message": "گزارش با موفقیت تولید شد"
                },
                status=status.HTTP_200_OK
            )
        except Exception as e:
            return Response(
                {
                    "status": "error",
                    "errors": {"error": str(e)},
                    "message": "خطا در تولید گزارش"
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class LaboratoryOperatorExcelReportAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        try:
            lab_id = request.query_params.get('lab_id')
            start_date = request.query_params.get('start_date')
            end_date = request.query_params.get('end_date')

            try:
                if start_date:
                    start_date = make_aware(datetime.datetime.strptime(start_date, '%Y-%m-%d'))
                if end_date:
                    end_date = make_aware(datetime.datetime.strptime(end_date, '%Y-%m-%d') + datetime.timedelta(days=1))
            except Exception:
                return Response(
                    {'error': 'Invalid date format. Use YYYY-MM-DD'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            labs = Laboratory.objects.all()
            if lab_id:
                labs = labs.filter(id=lab_id)

            def to_jalali(date_obj):
                if not date_obj:
                    return None
                if isinstance(date_obj, datetime.datetime):
                    date_obj = date_obj.date()
                return jdatetime.date.fromgregorian(date=date_obj).strftime('%Y/%m/%d')

            data = []
            for lab in labs:
                b_requests = Request.objects.filter(
                    experiment__laboratory=lab,
                    is_completed=True,
                    parent_request__isnull=True,
                    request_status__step__id__exact=8
                ).distinct()

                if start_date:
                    b_requests = b_requests.filter(
                        request_status__step__name='تکمیل شده',
                        request_status__created_at__gte=start_date
                    )
                if end_date:
                    b_requests = b_requests.filter(
                        request_status__step__name='تکمیل شده',
                        request_status__created_at__lt=end_date
                    )

                for req in b_requests:
                    payment_record = None
                    pr_qs = req.get_latest_order_payment_records()
                    transaction_amount = 0
                    refund_amount = 0
                    if pr_qs:
                        payment_record = pr_qs.filter(successful=True).order_by('created_at').first()
                        transaction_amount = pr_qs.filter(successful=True).aggregate(total=Sum('amount'))['total'] or 0
                        refund_amount = pr_qs.filter(is_returned=True).aggregate(total=Sum('amount'))['total'] or 0

                    completed_status = req.request_status.filter(
                        step__name='تکمیل شده'
                    ).order_by('-created_at').first()

                    child_discount = req.child_requests.exclude(request_status__step__name__in=['رد شده']).aggregate(
                        avg=Avg('discount'))['avg']
                    if child_discount is None:
                        child_discount = req.discount
                    completed_status = req.request_status.filter(step__name='تکمیل شده').first()
                    approver_name = (
                        completed_status.action_by.get_full_name() if completed_status and completed_status.action_by else None)

                    latest_result = req.request_results.order_by('-created_at').first()
                    if not latest_result:
                        latest_result = RequestResult.objects.filter(
                            request__in=req.child_requests.all()
                        ).order_by('-created_at').first()
                    operator_name = latest_result.result_by.get_full_name() if latest_result and latest_result.result_by else ''

                    Price_Before_Discount = req.price_wod / (1 - Decimal(child_discount or 0) / 100)
                    data.append({
                        'نام آزمایشگاه': lab.name,
                        'اپراتور': operator_name,
                        'شماره درخواست': req.request_number,
                        "وضعیت": req.lastest_status().step.name if req.lastest_status() else "",
                        'نام مشتری': req.owner.get_full_name(),
                        'تاریخ درخواست': to_jalali(req.created_at),
                        # 'تاریخ پرداخت': to_jalali(payment_record.created_at) if payment_record else None,
                        'تاریخ تکمیل درخواست': to_jalali(completed_status.created_at) if completed_status else None,
                        'تعداد نمونه': req.get_child_form_responses_count(),
                        'تعداد آزمون': req.child_requests.exclude(request_status__step__name__in=['رد شده']).count(),
                        'هزینه آزمون ها': Price_Before_Discount,
                        'درصد تخفیف': child_discount,
                        'مبلغ تخفیف': Price_Before_Discount * Decimal(child_discount or 0) / 100,
                        'گرنت لبزنت': req.labsnet_discount,
                        'گرنت پژوهشی': req.grant_request_discount,
                        'هزینه نهایی': req.price,
                        'مجموع پرداختی': transaction_amount,
                        'استرداد': abs(refund_amount),
                        # 'تکمیل توسط': approver_name,
                    })

            df = pd.DataFrame(data)
            df.replace('', pd.NA, inplace=True)
            df.dropna(axis=1, how='all', inplace=True)
            zero_cols = [col for col in df.select_dtypes(include='number').columns if (df[col].fillna(0) == 0).all()]
            df.drop(columns=zero_cols, inplace=True)

            now_str = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f'laboratory_operator_report_{now_str}.xlsx'
            full_path = os.path.join(settings.MEDIA_ROOT, 'reports', filename)
            os.makedirs(os.path.dirname(full_path), exist_ok=True)

            currency_cols = [
                'هزینه آزمون ها',
                'هزینه نهایی',
                'گرنت لبزنت',
                'گرنت پژوهشی',
                'مجموع پرداختی',
                'استرداد',
                'مبلغ تخفیف',
            ]
            for col_name in currency_cols:
                if col_name in df.columns:
                    df[col_name] = pd.to_numeric(df[col_name], errors='coerce')

            # add a totals row for numeric columns
            if not df.empty:
                numeric_cols = df.select_dtypes(include='number').columns
                totals = {col: df[col].sum() if col in numeric_cols else '' for col in df.columns}
                totals[df.columns[0]] = 'جمع'
                df = pd.concat([df, pd.DataFrame([totals])], ignore_index=True)

            with pd.ExcelWriter(full_path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='گزارش اپراتورها')

                worksheet = writer.sheets['گزارش اپراتورها']
                worksheet.auto_filter.ref = worksheet.dimensions

                currency_format = '#,##0 "ریال"'
                for col_name in currency_cols:
                    if col_name in df.columns:
                        col_idx = df.columns.get_loc(col_name) + 1
                        col_letter = get_column_letter(col_idx)
                        for row in range(2, worksheet.max_row + 1):
                            worksheet[f'{col_letter}{row}'].number_format = currency_format

            relative_path = os.path.relpath(full_path, settings.MEDIA_ROOT)
            file_url = request.build_absolute_uri(
                os.path.join(settings.MEDIA_URL, relative_path).replace('\\', '/')
            )

            return Response(
                {
                    "status": "success",
                    "data": {"url": file_url},
                    "message": "گزارش اپراتورها با موفقیت تولید شد"
                },
                status=status.HTTP_200_OK
            )
        except Exception as e:
            return Response(
                {
                    "status": "error",
                    "errors": {"error": str(e)},
                    "message": "خطا در تولید گزارش"
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class UserStatsAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        stats = User.objects.aggregate(
            total_staff=Count('id', filter=Q(user_type='staff')),
            total_customers=Count('id', filter=Q(user_type='customer')),
            partner_customers=Count('id', filter=Q(is_partner=True)),
            personal_customers=Count('id', filter=Q(user_type='customer', account_type='personal')),
            business_customers=Count('id', filter=Q(user_type='customer', account_type='business')),
        )

        return Response({
            "total_staff": stats["total_staff"],
            "total_customers": stats["total_customers"],
            "partner_customers": stats["partner_customers"],
            "personal_customers": stats["personal_customers"],
            "business_customers": stats["business_customers"],
        })


class LaboratoryStatsAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        lab_stats = Laboratory.objects.aggregate(
            total_active=Count('id', filter=Q(status='active')),
            total_inactive=Count('id', filter=~Q(status='active')),
            has_iso=Count('id', filter=Q(has_iso_17025=True)),
            no_iso=Count('id', filter=Q(has_iso_17025=False)),
        )

        device_stats = Device.objects.aggregate(
            active_devices=Count('id', filter=Q(status='active')),
            inactive_devices=Count('id', filter=~Q(status='active')),
        )

        experiment_stats = Experiment.objects.aggregate(
            active_experiments=Count('id', filter=Q(status='active')),
            inactive_experiments=Count('id', filter=~Q(status='active')),
        )

        return Response({
            "total_active": lab_stats["total_active"],
            "total_inactive": lab_stats["total_inactive"],
            "has_iso": lab_stats["has_iso"],
            "no_iso": lab_stats["no_iso"],
            "active_devices": device_stats["active_devices"],
            "inactive_devices": device_stats["inactive_devices"],
            "active_experiments": experiment_stats["active_experiments"],
            "inactive_experiments": experiment_stats["inactive_experiments"],
        })