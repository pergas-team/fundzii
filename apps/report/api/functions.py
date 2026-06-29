import os

import os
from datetime import datetime

from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import jdatetime
import pandas as pd
from django.db import connection
from django.db.models import Prefetch, Sum
from django.utils.timezone import is_aware
from apps.account.models import GrantTransaction
from apps.lab.models import Request
from django.conf import settings
from apps.order.models import PaymentRecord

def to_jalali(date_obj):
    if not date_obj:
        return None
    if isinstance(date_obj, datetime.datetime):
        date_obj = date_obj.date()
    return jdatetime.date.fromgregorian(date=date_obj).strftime('%Y/%m/%d')


def generate_excel_report():
    start_queries = len(connection.queries)
    data = []

    # Fetch all parent requests with related objects to minimize DB queries
    parent_requests = (
        Request.objects.filter(has_parent_request=False, is_completed=True)
            .select_related(
            'owner',
            'experiment',
            'experiment__laboratory',
            'experiment__device',
            'experiment__operator',
            'experiment__laboratory__technical_manager',
            'grant_request1',
            'grant_request2',
        )
            .prefetch_related(
            'parameter',
            Prefetch(
                'child_requests',
                queryset=Request.objects.select_related(
                    'experiment',
                    'experiment__laboratory',
                    'experiment__device',
                    'parent_request',
                ).prefetch_related('parameter'),
            ),
        )
            .annotate(
            grant_total=Sum('owner__received_records__granttransaction__amount'),
            payment_total=Sum('orders__order_payment_records__amount'),
        )
    )

    for parent_request in parent_requests:

        # Gather parent request data
        parent_data = {
            'شماره درخواست': parent_request.request_number,
            'نام مشتری': parent_request.owner.get_full_name(),
            'نام آزمون': parent_request.experiment.name,
            'مبلغ کل درخواست': parent_request.price,
            'تخفیف': parent_request.discount,
            'تاریخ ثبت درخواست': parent_request.created_at.replace(tzinfo=None) if is_aware(
                parent_request.created_at) else parent_request.created_at,
            "نام سازمان": parent_request.owner.company_name if parent_request.owner.account_type == 'business' else "",
            "شخصیت مشتری": parent_request.owner.account_type,
            "نوع مشتری": "دانشجو" if parent_request.owner.is_sharif_student else "سایر",
            "نام مشتری/نماینده قانونی": parent_request.owner.first_name,
            "نام خانوادگی مشتری/نماینده قانونی": parent_request.owner.last_name,
            "کدملی مشتری/نماینده قانونی": parent_request.owner.national_id,
            "شناسه ملی سازمان": parent_request.owner.company_national_id if parent_request.owner.account_type == 'business' else "",
            "کد اقتصادی سازمان مشتری": parent_request.owner.company_economic_number if parent_request.owner.account_type == 'business' else "",
            "شماره همراه مشتری/نماینده قانونی": parent_request.owner.username,
            "ایمیل مشتری": parent_request.owner.email,
            "نام آزمایشگاه": parent_request.experiment.laboratory.name,
            "کد کنترلی آزمایشگاه": parent_request.experiment.laboratory.control_code,
            "نام و نام اختصاری نوع دستگاه": parent_request.experiment.device.name,
            "برند و مدل دستگاه": parent_request.experiment.device.model,
            "کد کنترلی دستگاه": parent_request.experiment.device.control_code,
            "کد کنترلی آزمون": parent_request.experiment.control_code,
            "نام استاندارد": "استاندارد مربوطه",  # مقدار پیش‌فرض
            "کد کنترلی استاندارد": "کد استاندارد",  # مقدار پیش‌فرض
            "نام پارامتر": "پارامتر مربوطه",  # مقدار پیش‌فرض
            "نام اپراتور": parent_request.experiment.operator.first_name if parent_request.experiment.operator else "",
            "نام مدیر آزمایشگاه": parent_request.experiment.laboratory.technical_manager.first_name if parent_request.experiment.laboratory.technical_manager else "",
            "تعداد نمونه": parent_request.parameter.count(),
            "تاریخ دریافت نمونه": parent_request.created_at.date() if parent_request.created_at else "",
            "نوع واحد آزمون(مبنای تعیین تعرفه)": parent_request.experiment.test_unit_type,
            "تعرفه پارامتر": "",  # مقدار پیش‌فرض
            "هزینه کل آزمون": parent_request.price_wod,
            "درصد تخفیف": parent_request.discount,
            "مبلغ تخفیف": parent_request.price_wod - parent_request.price if parent_request.price_wod and parent_request.price else 0,
            "هزینه کل آزمون بعد از کسر تخفیف": parent_request.price,
            "نوع گرنت 1": parent_request.grant_request1 if parent_request.grant_request1 else "",
            "مبلغ گرنت 1": parent_request.grant_request1.approved_amount if parent_request.grant_request1 else 0,
            "نوع گرنت 2": parent_request.grant_request2 if parent_request.grant_request2 else "",
            "مبلغ گرنت 2": parent_request.grant_request2.approved_amount if parent_request.grant_request2 else 0,
            "مجموع مبلغ گرنت": sum([
                parent_request.grant_request1.approved_amount if parent_request.grant_request1 else 0,
                parent_request.grant_request2.approved_amount if parent_request.grant_request2 else 0
            ]),
            # "مبلغ کل پرداختی": parent_request.price - sum([
            #     parent_request.grant_request1.approved_amount if parent_request.grant_request1 else 0,
            #     parent_request.grant_request2.approved_amount if parent_request.grant_request2 else 0
            # ]),
            "مبلغ کل پرداختی": parent_request.price,
            "تاریخ ثبت موقت درخواست": parent_request.created_at.date() if parent_request.created_at else "",
            "تاریخ ثبت نهایی درخواست": parent_request.updated_at.date() if parent_request.updated_at else "",
            "تاریخ تعهد تحویل به مشتری": parent_request.delivery_date,
            "تاریخ انجام آزمون توسط اپراتور": "",  # مقدار پیش‌فرض
            "تاریخ تایید مدیر آزمایشگاه": "",  # مقدار پیش‌فرض
            "مهلت زمانی توقف به روز(از تاریخ ثبت اولیه درخواست تا تاریخ ثبت نهایی)": "",
            # (parent_request.updated_at - parent_request.created_at).days if parent_request.updated_at and parent_request.created_at else 0,
            "مهلت زمانی مقرر به روز(از تاریخ ثبت نهایی تا تاریخ تعهد تحویل به مشتری)": "",
            # (parent_request.delivery_date - parent_request.updated_at.date()).days if parent_request.delivery_date and parent_request.updated_at else 0,
            "مدت زمان توقف به روز(از تاریخ ثبت نهایی تا تاریخ تایید مدیر آز)": 0,  # مقدار پیش‌فرض
            "مدت زمان تاخیر به روز (مدت زمان توقف از مهلت زمانی مقرر کسر شود)": 0,  # مقدار پیش‌فرض
            "وضعیت": parent_request.lastest_status().step.name if parent_request.lastest_status() else "",
            "نام هماهنگ کننده": parent_request.owner.first_name if parent_request.owner else "",
        }

        # Fetch child requests for the parent request  (already prefetched)
        child_requests = parent_request.child_requests.all()

        for child_request in child_requests:
            if child_request.labsnet_status == 1:
                labsnet_status = 'ثبت نشده'
            elif child_request.labsnet_status == 2:
                labsnet_status = 'ثبت موفق'
            elif child_request.labsnet_status == 3:
                labsnet_status = 'ثبت ناموفق'
            else:
                labsnet_status = 'نامشخص'

            # Gather child request data
            child_data = {
                'شماره آزمون': child_request.request_number,
                'نام مشتری آزمون': child_request.parent_request.owner.get_full_name(),
                'نام آزمون': child_request.experiment.name,
                'مبلغ آزمون': child_request.price,
                'درصد تخفیف آزمون': child_request.discount,
                'تاریخ ثبت آزمون': child_request.created_at.replace(tzinfo=None) if is_aware(
                    child_request.created_at) else child_request.created_at,
                "وضعیت لبزنت": labsnet_status,
            }
            # Combine parent and child data
            combined_data = {**parent_data, **child_data}

            # Fetch parameters for the child request (prefetched)
            parameters_list = [param.name for param in child_request.parameter.all()]
            combined_data['پارامترها'] = ", ".join(parameters_list)

            # Use pre-aggregated totals for the parent request
            combined_data['مجموع گرنت پژوهشی'] = parent_request.grant_total or 0
            combined_data['مجموع مبلغ تراکنش ها'] = parent_request.payment_total or 0

            # Add row to data
            data.append(combined_data)

    # Convert data to pandas DataFrame
    df = pd.DataFrame(data)
    df.replace('', pd.NA, inplace=True)
    df.dropna(axis=1, how='all', inplace=True)
    zero_cols = [col for col in df.select_dtypes(include='number').columns if (df[col].fillna(0) == 0).all()]
    df.drop(columns=zero_cols, inplace=True)

    # Ensure the report is saved in the media folder
    media_path = os.path.join(settings.MEDIA_ROOT, 'reports')
    os.makedirs(media_path, exist_ok=True)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    file_path = os.path.join(media_path, f'Comprehensive_LIMS_Report_{timestamp}.xlsx')

    # Save to Excel file and apply formatting
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)
        wb = writer.book
        ws = wb.active

        # Determine accurate dimensions (rows/columns)
        max_row, max_col = ws.max_row, ws.max_column
        table_ref = f"A1:{get_column_letter(max_col)}{max_row}"

        # Freeze the header row (fixed header)
        ws.freeze_panes = 'A2'

        # Create Excel Table to apply autofilter & banded rows automatically
        table = Table(displayName="ReportTable", ref=table_ref)

        # Set style for alternate band coloring
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        table.tableStyleInfo = style

        # Add table to worksheet (autofilter included automatically)
        ws.add_table(table)

    end_queries = len(connection.queries)
    print(f"generate_excel_report executed {end_queries - start_queries} queries.")
    return file_path

